from django.shortcuts import render, redirect
from .Assembly_views import calculate_assembly_price
from .Pack_views import calculate_pack_price
from .NCT_views import calculate_nct_price
from .Volt_views import calculate_volt_price
from .MaterialCost_views import calculate_materialcost_price
from .Paint_views import calculate_paint_price
from .Jab_views import calculate_jab_price
from .Motor_views import calculate_motor_price
from .Controller_views import calculate_controller_price
from .Fan_views import calculate_fan_price
from .Bellmouth_views import calculate_bellmouth_price

import io
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from django.http import HttpResponse


자재비 = "자재비 (AL, SPCC 외)"
도장비 = "도장비"
NCT_가공비 = "NCT 가공비"
MOTOR = "MOTOR"
컨트롤러 = "컨트롤러"
FAN = "FAN"
BELLMOUTH = "BELLMOUTH (보호망포함)"
볼트 = "볼트&리벳"
포장용잡자재 = "포장용 잡자재"
조립인건비 = "조립인건비"
포장팔렛트 = "포장팔렛트"
운반비 = "운반비"
일반사양 = "일반사양"
고사양 = "고사양"

# OUTPUT 항목과 연관된 변수
def get_price_for_item(item_name, size, spec, motortype=None, motor_company=None, watt=None, **kwargs):
    if item_name == 조립인건비:
        return calculate_assembly_price(size, spec)

    elif item_name == 포장팔렛트:
        return calculate_pack_price(size, spec)

    elif item_name == NCT_가공비:
        return calculate_nct_price(size, spec)

    elif item_name == 볼트:
        return calculate_volt_price(size, spec)

    elif item_name == 자재비:
        return calculate_materialcost_price(size, spec)

    elif item_name == 도장비:
        return calculate_paint_price(size, spec)

    elif item_name == 포장용잡자재:
        return calculate_jab_price(size, spec)
        
    elif item_name == MOTOR:
        return calculate_motor_price(size, spec, motortype=motortype, motor_company=motor_company, watt=watt)
    
    elif item_name == 컨트롤러:
        return calculate_controller_price(size, spec, motortype=motortype, motor_company=motor_company, watt=watt)

    elif item_name == FAN:
        return calculate_fan_price(size, spec, motortype=motortype)
        
    elif item_name == BELLMOUTH:
        return calculate_bellmouth_price(size, spec, motortype=motortype)
    else:
        return 0

# INPUT에서의 변수
def ffuInput(request):
    if request.method == 'POST':
        size = request.POST.get('size') or 'Unknown Size'
        spec = request.POST.get('spec')
        motortype = request.POST.get('motortype')
        motor_company = request.POST.get('motor_company')
        watt = request.POST.get('watt')
        businessOwner = request.POST.get('businessOwner')
        location = request.POST.get('location')

        request.session['size'] = size
        request.session['businessOwner'] = businessOwner
        request.session['motortype'] = motortype
        request.session['location'] = location

        quantity_str = request.POST.get('quantity', '1')
        try:
            quantity = int(quantity_str)
        except ValueError:
            quantity = 1

        names = [자재비, 도장비, NCT_가공비, MOTOR, 컨트롤러, FAN, BELLMOUTH, 볼트, 포장용잡자재, 조립인건비, 포장팔렛트, 운반비]

        items = []
        for name in names:

            try:
                if name in [MOTOR, 컨트롤러]:
                    unit_price = get_price_for_item(name, size, spec, motortype=motortype, motor_company=motor_company, watt=watt)
                elif name in [FAN, BELLMOUTH]:
                    unit_price = get_price_for_item(name, size, spec, motortype=motortype)
                else:
                    unit_price = get_price_for_item(name, size, spec) 

                total_price = unit_price * quantity                         # 합계 = 단가 * 수량
                items.append((name, quantity, unit_price, total_price))     # 품명, 수량, 단가, 합계
            except Exception as e:
                print(f"Error occurred at {name}: {e}")  # 로깅을 위한 출력

        subtotal_unit = sum([item[2] for item in items])                    # UNIT 소계 단가
        subtotal_totals = sum([item[3] for item in items])                  # UNIT 소계 합계
        request.session['items'] = items

        context = {
            'items': items,
            'subtotal_unit': subtotal_unit,
            'subtotal_totals': subtotal_totals
        }

        request.session['subtotal_unit'] = subtotal_unit
        request.session['subtotal_totals'] = subtotal_totals

        return render(request, 'FFUOutput.html', context)

    return render(request, 'FFUInput.html')


# data to excel
def export_to_excel(items, businessOwner, motortype, location, size, subtotal_unit, subtotal_totals):
    wb = Workbook()
    ws = wb.active

    # 엑셀에서 26R*5C까지만 나타나게 설정
    for i in range(26, ws.max_row + 1):
        ws.delete_rows(i)
    for i in range(6, ws.max_column + 1):
        ws.delete_columns(i)

    # 배경색 지정
    colors = {
        "B5:E6": "DDD9C4", 
        "B19:E19": "D8E4BC",  
        "B21:E21": "FFFF00",  
        "B22:E22": "FABF8F",   
        "B25:E25": "C5D9F1"
    }

    for range_string, color in colors.items():
        for row in ws[range_string]:
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # 폰트 "맑은 고딕"으로 지정
    font = Font(name='맑은 고딕')

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    
    # 셀 병합 및 병합위치에 내용넣기
    ws.merge_cells('B5:B6')
    ws['B5'] = "품명"
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C5:E5')
    ws['C5'] = f"FFU {size}"
    ws['C5'].font = Font(bold=True)
    ws['C5'].alignment = Alignment(horizontal='center', vertical='center')

    # 열 너비 설정
    ws.column_dimensions['A'].width = 1.1
    ws.column_dimensions['B'].width = 26.9
    ws.column_dimensions['C'].width = 6.0
    ws.column_dimensions['D'].width = 11.8
    ws.column_dimensions['E'].width = 15.8

    # 행 높이 설정
    ws.row_dimensions[1].height = 9.6
    ws.row_dimensions[2].height = 30.6
    ws.row_dimensions[3].height = 19.8
    ws.row_dimensions[4].height = 19.8
    ws.row_dimensions[5].height = 30.6
    ws.row_dimensions[6].height = 25.2

    for i in range(7, 27):  # 7행부터 26행까지 높이 24.6
        ws.row_dimensions[i].height = 24.6

    # FFUInput에서 값 불러오기 위해 선언
    businessOwner = businessOwner
    motortype = motortype 
    location = location 
    size = size

    # 현장명, 유형, 지역
    alignment_center_left = Alignment(horizontal="left", vertical="center")
    alignment_center_center = Alignment(horizontal="center", vertical="center")
    alignment_center_right = Alignment(horizontal="right", vertical="center")
    alignment_center = Alignment(vertical="center")

    ws['B2'] = f'▣ 현장명: {businessOwner} FFU'
    ws['B2'].font = Font(bold=True, size = 14)
    ws['B2'].alignment = alignment_center_left

    ws['B3'] = f'▣ 유형: {motortype}'
    ws['B3'].font = Font(size = 14)
    ws['B3'].alignment = alignment_center_left

    ws['B4'] = f'▣ 지역: {location}'
    ws['B4'].font = Font(size = 14)
    ws['B4'].alignment = alignment_center_left

    # 품명, 수량, 단가, 합계의 제목 행
    font_11 = Font(size=11)

    start_row = 6
    cell_to_check = ws.cell(row=start_row, column=2)

    for range_ in ws.merged_cells.ranges:
        if cell_to_check.coordinate in range_:
            cell = ws[range_.start_cell.coordinate]
            cell.value = "품명"
            cell.font = font_11  
            break
    else:
        cell_to_check.value = "품명"
        cell_to_check.font = font_11  

    cell_quantity = ws.cell(row=start_row, column=3, value="수량")
    cell_quantity.font = font_11  

    cell_price = ws.cell(row=start_row, column=4, value="단가")
    cell_price.font = font_11  

    cell_total = ws.cell(row=start_row, column=5, value="합계")
    cell_total.font = font_11

    for cell in ws["C6":"E6"]:
        for c in cell:
            c.alignment = alignment_center_center

    for cell in ws["C7":"C25"]:
        for c in cell:
            c.alignment = alignment_center_center

    for cell in ws["D7":"E25"]:
        for c in cell:
            c.alignment = alignment_center_right
    
    for cell_addr in ["B19", "B21", "B22", "B25", "B26"]:
        ws[cell_addr].alignment = alignment_center_center

    cells_to_center = list(ws["B7":"B18"]) + [(ws["B20"],), (ws["B23"],), (ws["B24"],)]

    for cell_tuple in cells_to_center:
        for cell in cell_tuple:
            cell.alignment = alignment_center


    # 정수 표기, 1,000으로 표현
    comma_style = NamedStyle(name="comma_style", number_format="#,##0")

    if 'comma_style' not in wb.named_styles:
        wb.add_named_style(comma_style)

    for row in ws.iter_rows(min_row=6, max_row=25, min_col=3, max_col=5):  # C6 to E25
        for cell in row:
            try:
                value = float(cell.value)
                
                if value.is_integer():
                    cell.value = int(value)
                
                cell.style = 'comma_style'
            except (ValueError, TypeError):
                pass



    # 품목 데이터 추가
    for i, item in enumerate(items, start=start_row+1):  # 아이템 시작을 7행부터
        for j, value in enumerate(item, start=2):  # B열부터 시작
            ws.cell(row=i, column=j, value=value)

    # UNIT 소계
    ws["B19"].value = "UNIT 소계"
    ws['D19'] = f'\\   {int(subtotal_unit):,}'
    ws['D19'].number_format = "#,##0"  

    ws['E19'] = f'\\   {int(subtotal_totals):,}'
    ws['E19'].number_format = "#,##0"

    # FILTER 소계, 원가 합계, 총계, 비고
    ws["B20"].value = "FILTER"  
    ws["B21"].value = "FILTER 소계"
    ws["B22"].value = "원가 합계"
    ws["B23"].value = "관리비"
    ws["B24"].value = "영업이익"
    ws["B25"].value = "총계"
    ws["B26"].value = "비고"

    ws.merge_cells('C26:E26')
    for cell in ws["C26":"E26"]:
        for c in cell:
            c.alignment = Alignment(horizontal='center')

    # B7부터 E18까지 글씨크기 10
    font_size_10 = Font(size=10)
    for row in ws.iter_rows(min_row=7, max_row=18, min_col=2, max_col=5):
        for cell in row:
            cell.font = font_size_10

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_side = Side(style='medium')

    ws = ws 

    # 1. B7부터 E25까지 모든 테두리
    for row in ws.iter_rows(min_row=7, max_row=25, min_col=2, max_col=5):
        for cell in row:
            cell.border = thin_border

    # 2. B26은 오른쪽 테두리
    cell = ws["B26"]
    cell.border = Border(right=thin_border.right)


    # 3. C6부터 E6까지 모든 테두리
    for cell in ws["C6":"E6"]:
        for c in cell:
            c.border = thin_border

    # 4. B5부터 B6까지 오른쪽 테두리
    for cell in ws["B5":"B6"]:
        for c in cell:
            c.border = Border(right=thin_border.right)

    # 5. B5부터 E26까지 굵은 바깥쪽 테두리
    for row in ws.iter_rows(min_row=5, max_row=26, min_col=2, max_col=5):
        for cell in row:
            # 최상단 행이나 최하단 행에 대한 상하측 테두리 설정
            top_border = thick_side if cell.row == 5 else cell.border.top
            bottom_border = thick_side if cell.row == 26 else cell.border.bottom

            # 왼쪽 끝 열이나 오른쪽 끝 열에 대한 좌우측 테두리 설정
            left_border = thick_side if cell.column == 2 else cell.border.left
            right_border = thick_side if cell.column == 5 else cell.border.right

            cell.border = Border(top=top_border, bottom=bottom_border, left=left_border, right=right_border)


    # 6 ~ 10. 굵은 바깥쪽 테두리
    for cell_range in ["B5:E6", "B19:E19", "B21:E21", "B22:E22", "B25:E25"]:
        start_col = ord(cell_range.split(':')[0][0]) - 65
        end_col = ord(cell_range.split(':')[1][0]) - 65
        start_row = int(cell_range.split(':')[0][1:])
        end_row = int(cell_range.split(':')[1][1:])
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col + 1, max_col=end_col + 1):
            for cell in row:
                top_border = thick_side if cell.row == start_row else cell.border.top
                bottom_border = thick_side if cell.row == end_row else cell.border.bottom
                left_border = thick_side if cell.column == start_col + 1 else cell.border.left
                right_border = thick_side if cell.column == end_col + 1 else cell.border.right

                cell.border = Border(top=top_border, bottom=bottom_border, left=left_border, right=right_border)

    return wb

# excel 다운로드
def download_excel(request):
    items = request.session.get('items', [])  
    businessOwner = request.session.get('businessOwner', '')
    motortype = request.session.get('motortype', '')
    location = request.session.get('location', '')
    size = request.session.get('size', 'Unknown Size')
    subtotal_unit = request.session.get('subtotal_unit', '')
    subtotal_totals = request.session.get('subtotal_totals', '')

    wb = export_to_excel(items, businessOwner, motortype, location, size, subtotal_unit, subtotal_totals)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="FFU.xlsx"'


    return response

def ffuOutput(request):
    return render(request, 'FFUOutput.html')

def ffuDashboard(request):
    return render(request, 'FFUDashboard.html')    
